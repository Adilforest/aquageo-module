"""Condition-summary report export tests (issue #31), PostGIS.

The report must (a) stream a real PDF and a real .xlsx with the right content
types, (b) carry numbers identical to the /api/v1/stats/* aggregates it reuses,
and (c) react to the same map filters.
"""
import io

import pytest
from django.contrib.gis.geos import Point
from openpyxl import load_workbook
from rest_framework.test import APIClient

from catalog.models import AdminUnit, Basin, ObjectType, Structure

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


@pytest.fixture
def world(db):
    dam = ObjectType.objects.create(code="dam", name_ru="Плотина", geometry_kind="point")
    canal = ObjectType.objects.create(code="canal", name_ru="Канал", geometry_kind="line")
    basin_a = Basin.objects.create(name_ru="Шу-Таласский")
    basin_b = Basin.objects.create(name_ru="Иной")
    d1 = AdminUnit.objects.create(kato="3110", name_ru="Байзакский", level="district")

    def mk(t, name, cond, basin=basin_a, district=d1):
        return Structure.objects.create(
            type=t, name_ru=name, condition_status=cond, basin=basin, admin_unit=district,
            geom=Point(71.0, 43.0, srid=4326), attributes={},
        )

    mk(dam, "d1", "emergency")
    mk(dam, "d2", "repair")
    mk(canal, "c1", "serviceable", basin=basin_b)
    mk(canal, "c2", "monitoring")
    return {"basin_a": basin_a, "basin_b": basin_b}


URL = "/api/v1/reports/condition-summary/"


@pytest.mark.django_db
def test_pdf_export_is_a_nonempty_pdf(world):
    resp = APIClient().get(URL, {"format": "pdf"})
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/pdf"
    assert "attachment" in resp["Content-Disposition"]
    body = resp.getvalue()
    assert len(body) > 500
    assert body[:5] == b"%PDF-"


@pytest.mark.django_db
def test_pdf_is_the_default_format(world):
    resp = APIClient().get(URL)
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/pdf"
    assert resp.getvalue()[:5] == b"%PDF-"


@pytest.mark.django_db
def test_xlsx_export_is_a_nonempty_workbook(world):
    resp = APIClient().get(URL, {"format": "xlsx"})
    assert resp.status_code == 200
    assert resp["Content-Type"] == XLSX_CONTENT_TYPE
    assert "condition-summary.xlsx" in resp["Content-Disposition"]
    body = resp.getvalue()
    assert len(body) > 500
    # A real .xlsx is a ZIP container — check the PK magic and that it opens.
    assert body[:2] == b"PK"
    wb = load_workbook(io.BytesIO(body))
    assert "Состояние" in wb.sheetnames


def _condition_counts_from_xlsx(body):
    """Read the condition distribution back out of the first sheet."""
    wb = load_workbook(io.BytesIO(body))
    ws = wb["Состояние"]
    label_to_code = {
        "Исправное": "serviceable",
        "Требует наблюдения": "monitoring",
        "Требует ремонта": "repair",
        "Аварийное": "emergency",
    }
    counts = {}
    for row in ws.iter_rows(min_row=9, max_row=12, values_only=True):
        label, value = row[0], row[1]
        if label in label_to_code:
            counts[label_to_code[label]] = value
    total = ws["B5"].value
    index = ws["B6"].value
    return counts, total, index


@pytest.mark.django_db
def test_report_numbers_match_stats_aggregates(world):
    client = APIClient()
    stats = client.get("/api/v1/stats/by-condition/").data

    body = client.get(URL, {"format": "xlsx"}).getvalue()
    counts, total, index = _condition_counts_from_xlsx(body)

    assert counts == stats["counts"]
    assert total == stats["total"]
    assert index == stats["index"]
    # Sanity against the fixture: one of each condition, index 50.
    assert total == 4
    assert index == 50


@pytest.mark.django_db
def test_by_type_sheet_matches_stats(world):
    client = APIClient()
    stats = {r["type"]: r["count"] for r in client.get("/api/v1/stats/by-type/").data}

    body = client.get(URL, {"format": "xlsx"}).getvalue()
    wb = load_workbook(io.BytesIO(body))
    ws = wb["По типам"]
    from_report = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            from_report[row[1]] = row[2]
    assert from_report == stats == {"dam": 2, "canal": 2}


@pytest.mark.django_db
def test_report_reacts_to_filters(world):
    client = APIClient()
    # Restrict to basin_b (one serviceable canal) — both stats and report agree.
    params = {"format": "xlsx", "basin": str(world["basin_b"].pk)}
    stats = client.get(
        "/api/v1/stats/by-condition/", {"basin": str(world["basin_b"].pk)}
    ).data
    body = client.get(URL, params).getvalue()
    counts, total, index = _condition_counts_from_xlsx(body)

    assert total == stats["total"] == 1
    assert counts["serviceable"] == 1
    assert counts["emergency"] == 0

    # Filtering by type=dam yields a different total than the unfiltered report.
    dam_total = _condition_counts_from_xlsx(
        client.get(URL, {"format": "xlsx", "type": "dam"}).getvalue()
    )[1]
    assert dam_total == 2


@pytest.mark.django_db
def test_pdf_reacts_to_filters_without_error(world):
    # Empty result set (no such basin) still renders a valid PDF.
    resp = APIClient().get(URL, {"format": "pdf", "type": "nonexistent"})
    assert resp.status_code == 200
    assert resp.getvalue()[:5] == b"%PDF-"
