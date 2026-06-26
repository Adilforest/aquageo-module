"""Condition-summary report: data assembly + PDF/Excel rendering (issue #31).

Every number comes straight from :mod:`catalog.stats_service` — the same source
the dashboard ``/api/v1/stats/*`` endpoints and the map use — so an exported
report can never drift from what the UI shows. This module only *formats* those
aggregates into a PDF or an ``.xlsx`` workbook; it does not recompute anything.
"""
import io

from django.utils import timezone

from catalog import stats_service
from catalog.models import ConditionStatus

# Condition colours — single source of truth across the stack. Keep in sync with
# frontend/src/theme.ts (CONDITION_COLORS). Codes are the canonical hex values.
CONDITION_COLORS = {
    "serviceable": "#1F9D57",
    "monitoring": "#2F73E0",
    "repair": "#E0921E",
    "emergency": "#E0443E",
    "nodata": "#8C97A7",
}

# Human labels for condition codes — reuse the model's TextChoices labels.
CONDITION_LABELS = {c.value: c.label for c in ConditionStatus}

RISK_LEVELS = ["critical", "high", "watch", "none"]
RISK_LABELS = {
    "critical": "Критический",
    "high": "Высокий",
    "watch": "Наблюдение",
    "none": "Нет",
}

# Filter query params reflected in the report header (same names as map/dashboard).
FILTER_KEYS = ["type", "condition", "basin", "district", "search"]


def collect_filters(request):
    """Active map filters as an ordered list of (label, value) for the header."""
    out = []
    for key in FILTER_KEYS:
        values = request.GET.getlist(key)
        if values:
            out.append((key, ", ".join(values)))
    return out


def build_condition_summary(request):
    """Assemble the full condition-summary payload from the shared stats service."""
    qs = stats_service.filtered_structures(request)
    return {
        "generated_at": timezone.now(),
        "filters": collect_filters(request),
        "condition": stats_service.by_condition(qs),
        "by_type": stats_service.by_type(qs),
        "by_territory": stats_service.by_territory(qs, "basin"),
        "risk": stats_service.risk_summary(qs),
    }


TITLE = "Сводный отчёт о состоянии инфраструктуры"


def _filters_line(report):
    if not report["filters"]:
        return "Фильтры: не заданы (все объекты)"
    parts = [f"{k}={v}" for k, v in report["filters"]]
    return "Фильтры: " + "; ".join(parts)


# --------------------------------------------------------------------------- #
# Excel (openpyxl)
# --------------------------------------------------------------------------- #
def render_xlsx(report):
    """Render the report as an .xlsx workbook (bytes)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    # Sheet 1: summary + condition distribution.
    ws = wb.active
    ws.title = "Состояние"
    bold = Font(bold=True)

    ws["A1"] = TITLE
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Сформирован: " + report["generated_at"].strftime("%Y-%m-%d %H:%M")
    ws["A3"] = _filters_line(report)

    cond = report["condition"]
    ws["A5"] = "Всего объектов"
    ws["B5"] = cond["total"]
    ws["A6"] = "Индекс состояния (0–100)"
    ws["B6"] = cond["index"]

    ws["A8"] = "Состояние"
    ws["B8"] = "Количество"
    ws["A8"].font = bold
    ws["B8"].font = bold
    row = 9
    for code in ConditionStatus.values:
        ws.cell(row=row, column=1, value=CONDITION_LABELS.get(code, code))
        ws.cell(row=row, column=2, value=cond["counts"].get(code, 0))
        fill = PatternFill(
            start_color=CONDITION_COLORS[code].lstrip("#"),
            end_color=CONDITION_COLORS[code].lstrip("#"),
            fill_type="solid",
        )
        ws.cell(row=row, column=1).fill = fill
        row += 1

    # Sheet 2: by type.
    ws2 = wb.create_sheet("По типам")
    ws2["A1"] = "Тип объекта"
    ws2["B1"] = "Код"
    ws2["C1"] = "Количество"
    for c in ("A1", "B1", "C1"):
        ws2[c].font = bold
    for i, r in enumerate(report["by_type"], start=2):
        ws2.cell(row=i, column=1, value=r["type_name"])
        ws2.cell(row=i, column=2, value=r["type"])
        ws2.cell(row=i, column=3, value=r["count"])

    # Sheet 3: by territory (basin).
    ws3 = wb.create_sheet("По бассейнам")
    ws3["A1"] = "Бассейн"
    ws3["B1"] = "Количество"
    ws3["A1"].font = bold
    ws3["B1"].font = bold
    for i, item in enumerate(report["by_territory"]["items"], start=2):
        ws3.cell(row=i, column=1, value=item["name"])
        ws3.cell(row=i, column=2, value=item["count"])

    # Sheet 4: risk summary.
    ws4 = wb.create_sheet("Риски")
    risk = report["risk"]
    ws4["A1"] = "Гидропостов учтено"
    ws4["B1"] = risk["hydroposts"]
    ws4["A2"] = "Прогноз пересекает опасный уровень"
    ws4["B2"] = risk["forecast_crossing"]
    ws4["A4"] = "Уровень риска"
    ws4["B4"] = "Паводок"
    ws4["C4"] = "Маловодье"
    for c in ("A4", "B4", "C4"):
        ws4[c].font = bold
    for i, level in enumerate(RISK_LEVELS, start=5):
        ws4.cell(row=i, column=1, value=RISK_LABELS[level])
        ws4.cell(row=i, column=2, value=risk["flood"].get(level, 0))
        ws4.cell(row=i, column=3, value=risk["low_water"].get(level, 0))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# PDF (reportlab)
# --------------------------------------------------------------------------- #
def render_pdf(report):
    """Render the report as a PDF document (bytes)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    styles = getSampleStyleSheet()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
        title=TITLE,
    )
    flow = []

    flow.append(Paragraph(TITLE, styles["Title"]))
    flow.append(Paragraph(
        "Сформирован: " + report["generated_at"].strftime("%Y-%m-%d %H:%M"),
        styles["Normal"],
    ))
    flow.append(Paragraph(_filters_line(report), styles["Normal"]))
    flow.append(Spacer(1, 8 * mm))

    cond = report["condition"]
    flow.append(Paragraph(
        f"Всего объектов: <b>{cond['total']}</b> &nbsp;·&nbsp; "
        f"Индекс состояния: <b>{cond['index']}</b>/100",
        styles["Normal"],
    ))
    flow.append(Spacer(1, 4 * mm))

    # Condition distribution table with the canonical condition colours.
    flow.append(Paragraph("Распределение по состоянию", styles["Heading2"]))
    data = [["Состояние", "Количество"]]
    for code in ConditionStatus.values:
        data.append([CONDITION_LABELS.get(code, code), cond["counts"].get(code, 0)])
    table = Table(data, colWidths=[90 * mm, 40 * mm])
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16324F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C8D0DA")),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
    ]
    for i, code in enumerate(ConditionStatus.values, start=1):
        style.append(("BACKGROUND", (0, i), (0, i), colors.HexColor(CONDITION_COLORS[code])))
        style.append(("TEXTCOLOR", (0, i), (0, i), colors.white))
    table.setStyle(TableStyle(style))
    flow.append(table)
    flow.append(Spacer(1, 6 * mm))

    # By type.
    flow.append(Paragraph("Распределение по типам", styles["Heading2"]))
    type_data = [["Тип объекта", "Количество"]]
    for r in report["by_type"]:
        type_data.append([r["type_name"] or r["type"], r["count"]])
    if len(type_data) == 1:
        type_data.append(["—", 0])
    t2 = Table(type_data, colWidths=[90 * mm, 40 * mm])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16324F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C8D0DA")),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
    ]))
    flow.append(t2)
    flow.append(Spacer(1, 6 * mm))

    # By territory.
    flow.append(Paragraph("Распределение по бассейнам", styles["Heading2"]))
    terr_data = [["Бассейн", "Количество"]]
    for item in report["by_territory"]["items"]:
        terr_data.append([item["name"], item["count"]])
    if len(terr_data) == 1:
        terr_data.append(["—", 0])
    t3 = Table(terr_data, colWidths=[90 * mm, 40 * mm])
    t3.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16324F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C8D0DA")),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
    ]))
    flow.append(t3)
    flow.append(Spacer(1, 6 * mm))

    # Risk summary.
    risk = report["risk"]
    flow.append(Paragraph("Риск-сводка (гидропосты)", styles["Heading2"]))
    flow.append(Paragraph(
        f"Гидропостов учтено: <b>{risk['hydroposts']}</b> &nbsp;·&nbsp; "
        f"прогноз пересекает опасный уровень: <b>{risk['forecast_crossing']}</b>",
        styles["Normal"],
    ))
    risk_data = [["Уровень риска", "Паводок", "Маловодье"]]
    for level in RISK_LEVELS:
        risk_data.append([
            RISK_LABELS[level], risk["flood"].get(level, 0), risk["low_water"].get(level, 0),
        ])
    t4 = Table(risk_data, colWidths=[60 * mm, 35 * mm, 35 * mm])
    t4.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16324F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C8D0DA")),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
    ]))
    flow.append(t4)

    doc.build(flow)
    return buf.getvalue()
